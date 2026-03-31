from openpyxl import load_workbook
import os
wb = load_workbook(os.path.expanduser('~/Downloads/CTI-CMM-v1.3.xlsx'), data_only=True)
ws = wb['Domain 7 - THIRD-PARTIES']
for row in range(1, 80):
    vals = [ws.cell(row=row, column=c).value for c in range(1, 10)]
    if any(v is not None for v in vals):
        b = repr(vals[1])[:30]
        c = str(vals[2])[:80] if vals[2] else 'None'
        print(f'Row {row}: B={b}  C={c}  D={vals[3]}  E={vals[4]}')
wb.close()
