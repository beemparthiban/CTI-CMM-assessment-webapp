#!/usr/bin/env python3
"""Extract CTI-CMM assessment data from Excel to JSON."""

import json
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook


EXCEL_PATH = os.path.expanduser("~/Downloads/CTI-CMM-v1.3.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "src", "data", "domains.json")

# Domain sheet names start with "Domain"
DOMAIN_SHEET_PREFIX = "Domain"


def extract_domains(wb):
    """Extract all domain data from the workbook."""
    domains = []
    
    # Get dashboard data for nicknames and domain names
    # Dashboard rows 6-16 map to domains 1-11
    dashboard = {}
    if "Dashboard" in wb.sheetnames:
        ds = wb["Dashboard"]
        for row in range(6, 17):  # rows 6-16
            name = ds.cell(row=row, column=2).value
            nickname = ds.cell(row=row, column=3).value
            if name and nickname:
                domain_id = row - 5  # row 6 → domain 1, row 7 → domain 2, etc.
                dashboard[domain_id] = {"name": str(name).strip(), "nickname": str(nickname).strip()}
    
    domain_sheets = [s for s in wb.sheetnames if s.startswith(DOMAIN_SHEET_PREFIX)]
    
    for sheet_name in domain_sheets:
        ws = wb[sheet_name]
        
        # Extract domain number from sheet name, e.g. "Domain 1 - ASSET"
        match = re.match(r"Domain\s+(\d+)", sheet_name)
        if not match:
            continue
        domain_id = int(match.group(1))
        
        # Get domain name from row 4, col B, or from dashboard
        domain_name_cell = ws.cell(row=4, column=2).value
        if domain_name_cell:
            # Remove domain number prefix and nickname suffix like "1. ... (ASSET)"
            domain_name = str(domain_name_cell).strip()
            domain_name = re.sub(r"^\d+\.\s*", "", domain_name)
            domain_name = re.sub(r"\s*\([^)]+\)\s*$", "", domain_name)
        elif domain_id in dashboard:
            # Dashboard name has "1. ..." prefix too
            dn = dashboard[domain_id]["name"]
            domain_name = re.sub(r"^\d+\.\s*", "", dn)
        else:
            domain_name = sheet_name
        
        # Get nickname from sheet tab name (e.g. "Domain 7 - THIRD-PARTIES" → "THIRD-PARTIES")
        # Sheet tab names are unique; dashboard nicknames may have duplicates (e.g. RISK)
        nickname_match = re.search(r"-\s*(.+)$", sheet_name)
        nickname = nickname_match.group(1).strip() if nickname_match else ""
        if not nickname and domain_id in dashboard:
            nickname = dashboard[domain_id].get("nickname", "")
        
        sections = []
        current_section = None
        current_maturity = None
        section_counter = 0
        obj_counter = 0
        
        # Scan rows starting from row 6 onward
        for row in range(6, ws.max_row + 1):
            col_b = ws.cell(row=row, column=2).value  # Maturity level or section name
            col_c = ws.cell(row=row, column=3).value  # Objective text
            col_d = ws.cell(row=row, column=4).value  # Score
            col_e = ws.cell(row=row, column=5).value  # Max
            
            # Skip empty rows
            if col_b is None and col_c is None:
                continue
            
            col_b_str = str(col_b).strip() if col_b else ""
            col_c_str = str(col_c).strip() if col_c else ""
            
            # Skip subtotal and total rows
            if col_c_str in ("Section Subtotal", "Domain Total"):
                continue
            
            # Check if this is a section header row
            # Section headers have content in col B but cols C-I are generally null
            # They typically have a numbered format like "1. Asset Visibility"
            is_section_header = False
            if col_b_str and not col_c_str:
                # Check if cols D, E, F are all None (section header pattern)
                col_f = ws.cell(row=row, column=6).value
                if col_d is None and col_e is None and col_f is None:
                    # This looks like a section header
                    is_section_header = True
            
            if is_section_header:
                section_counter += 1
                obj_counter = 0
                current_section = {
                    "id": f"{domain_id}.{section_counter}",
                    "name": col_b_str,
                    "objectives": []
                }
                sections.append(current_section)
                current_maturity = None  # Reset maturity for new section
                continue
            
            # Check if this is an objective row (has text in col C)
            if col_c_str:
                # Update maturity level if col B has a value
                if col_b_str and col_b_str.startswith("CTI"):
                    current_maturity = col_b_str
                
                # If no section yet, create a default one
                if current_section is None:
                    section_counter += 1
                    obj_counter = 0
                    current_section = {
                        "id": f"{domain_id}.{section_counter}",
                        "name": "General",
                        "objectives": []
                    }
                    sections.append(current_section)
                
                obj_counter += 1
                # Generate objective ID using letters
                obj_letter = chr(96 + obj_counter) if obj_counter <= 26 else f"a{obj_counter - 26}"
                
                max_score = col_e if col_e and isinstance(col_e, (int, float)) else 3
                
                objective = {
                    "id": f"{current_section['id']}.{obj_letter}",
                    "maturityLevel": current_maturity or "CTI1",
                    "text": col_c_str,
                    "maxScore": int(max_score)
                }
                current_section["objectives"].append(objective)
        
        # Calculate total max score
        total_max = sum(
            obj["maxScore"]
            for sec in sections
            for obj in sec["objectives"]
        )
        
        domain = {
            "id": domain_id,
            "name": domain_name,
            "nickname": nickname,
            "maxScore": total_max,
            "sections": sections
        }
        domains.append(domain)
    
    return domains


def main():
    print(f"Loading workbook from {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH, read_only=False, data_only=True)
    
    print(f"Sheets found: {wb.sheetnames}")
    
    domains = extract_domains(wb)
    
    # Print summary
    total_objectives = 0
    for d in domains:
        obj_count = sum(len(s["objectives"]) for s in d["sections"])
        total_objectives += obj_count
        print(f"  Domain {d['id']}: {d['nickname']} — {len(d['sections'])} sections, {obj_count} objectives, max={d['maxScore']}")
    
    print(f"\nTotal: {len(domains)} domains, {total_objectives} objectives")
    
    output = {"domains": domains}
    
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)
    
    print(f"\nWritten to {os.path.abspath(OUTPUT_PATH)}")
    wb.close()


if __name__ == "__main__":
    main()
